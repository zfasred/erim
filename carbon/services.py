# carbon/services.py
"""Karbon hesaplama ve veri yükleme servisleri"""

import pandas as pd
from decimal import Decimal
from datetime import date, datetime
from django.db import transaction
from django.db.models import Sum, Q

from carbon.models import (
    EmissionFactor, GWPFactor, 
    Scope1Data, Scope2Data, Scope3Data, Scope4Data,
    ProductCarbonAllocation, Report
)
from core.models import Firm, User


class ExcelDataLoader:
    """Excel'den emisyon faktörlerini ve sabit katsayıları yükle"""
    
    @staticmethod
    @transaction.atomic
    def load_emission_factors():
        """Sabit emisyon faktörlerini yükle"""
        
        # GWP değerlerini yükle
        GWPFactor.objects.update_or_create(
            gas='CH4',
            defaults={
                'value': Decimal('27.9'),
                'source': 'IPCC AR5',
                'valid_from': date(2014, 1, 1)
            }
        )
        
        GWPFactor.objects.update_or_create(
            gas='N2O',
            defaults={
                'value': Decimal('273'),
                'source': 'IPCC AR5',
                'valid_from': date(2014, 1, 1)
            }
        )
        
        # Yakıt emisyon faktörleri
        fuel_factors = [
            {
                'name': 'Doğalgaz',
                'category': 'KAPSAM_1',
                'factor_type': 'FUEL',
                'ef_co2': Decimal('56100'),
                'ef_ch4': Decimal('1'),
                'ef_n2o': Decimal('0.1'),
                'ncv': Decimal('48'),
                'density': Decimal('0.72'),
                'valid_from': date(2024, 1, 1),
                'source': 'IPCC 2006'
            },
            {
                'name': 'Motorin',
                'category': 'KAPSAM_1',
                'factor_type': 'FUEL',
                'ef_co2': Decimal('74100'),
                'ef_ch4': Decimal('3'),
                'ef_n2o': Decimal('0.6'),
                'ncv': Decimal('43'),
                'density': Decimal('0.835'),
                'valid_from': date(2024, 1, 1),
                'source': 'IPCC 2006'
            },
            {
                'name': 'Benzin',
                'category': 'KAPSAM_1',
                'factor_type': 'FUEL',
                'ef_co2': Decimal('69300'),
                'ef_ch4': Decimal('3'),
                'ef_n2o': Decimal('0.6'),
                'ncv': Decimal('44.3'),
                'density': Decimal('0.745'),
                'valid_from': date(2024, 1, 1),
                'source': 'IPCC 2006'
            }
        ]
        
        for factor_data in fuel_factors:
            EmissionFactor.objects.update_or_create(
                name=factor_data['name'],
                valid_from=factor_data['valid_from'],
                defaults=factor_data
            )
        
        # Elektrik emisyon faktörü
        EmissionFactor.objects.update_or_create(
            name='Türkiye Elektrik Şebekesi',
            valid_from=date(2024, 1, 1),
            defaults={
                'category': 'KAPSAM_2',
                'factor_type': 'ELECTRICITY',
                'ef_co2': Decimal('0.442'),  # tCO2/MWh olarak
                'ef_ch4': Decimal('0'),
                'ef_n2o': Decimal('0'),
                'value': 0.442,  # Backward compatibility
                'source': 'TEİAŞ 2024'
            }
        )
        
        # Malzeme emisyon faktörleri
        material_factors = [
            {'name': 'Çelik', 'ef_co2': Decimal('1.85'), 'category': 'KAPSAM_4'},
            {'name': 'Dökme Demir', 'ef_co2': Decimal('2.00'), 'category': 'KAPSAM_4'},
            {'name': 'Alüminyum', 'ef_co2': Decimal('8.24'), 'category': 'KAPSAM_4'},
            {'name': 'Bakır', 'ef_co2': Decimal('3.83'), 'category': 'KAPSAM_4'},
            {'name': 'Plastik (PP)', 'ef_co2': Decimal('1.90'), 'category': 'KAPSAM_4'},
            {'name': 'Magnezyum', 'ef_co2': Decimal('8.50'), 'category': 'KAPSAM_4'},
        ]
        
        for material in material_factors:
            EmissionFactor.objects.update_or_create(
                name=material['name'],
                valid_from=date(2024, 1, 1),
                defaults={
                    'category': material['category'],
                    'factor_type': 'MATERIAL',
                    'ef_co2': material['ef_co2'],
                    'value': float(material['ef_co2']),  # Backward compatibility
                    'source': 'Ecoinvent 3.8'
                }
            )
        
        print("✅ Emisyon faktörleri başarıyla yüklendi!")


class CarbonCalculationService:
    """Excel'deki hesaplama mantığını uygulayan servis"""
    
    def __init__(self, firm, year, month=None, user=None):
        self.firm = firm
        self.year = year
        self.month = month
        self.user = user
        self.report = None
    
    @transaction.atomic
    def load_example_data(self):
        """Excel'deki örnek verileri yükle"""
        
        # Kapsam 1 - Sabit Yanma (Doğalgaz)
        doğalgaz_factor = EmissionFactor.objects.get(
            name='Doğalgaz',
            valid_from__lte=date(self.year, self.month or 1, 1)
        )
        
        scope1_data = [
            {
                'location': 'ASANSÖR D2',
                'fuel_name': 'Doğalgaz',
                'consumption_value': Decimal('1751988'),
                'consumption_unit': 'm³',
                'emission_type': 'STATIONARY',
                'emission_factor': doğalgaz_factor
            },
            {
                'location': 'DÖKÜMHANE D3',
                'fuel_name': 'Doğalgaz',
                'consumption_value': Decimal('7118068'),
                'consumption_unit': 'm³',
                'emission_type': 'STATIONARY',
                'emission_factor': doğalgaz_factor
            },
            {
                'location': 'FRENBU',
                'fuel_name': 'Doğalgaz',
                'consumption_value': Decimal('326622'),
                'consumption_unit': 'm³',
                'emission_type': 'STATIONARY',
                'emission_factor': doğalgaz_factor
            }
        ]
        
        for data in scope1_data:
            Scope1Data.objects.update_or_create(
                firm=self.firm,
                location=data['location'],
                fuel_name=data['fuel_name'],
                period_year=self.year,
                period_month=self.month or 1,
                defaults={
                    'consumption_value': data['consumption_value'],
                    'consumption_unit': data['consumption_unit'],
                    'emission_type': data['emission_type'],
                    'emission_factor': data['emission_factor'],
                    'created_by': self.user
                }
            )
        
        # Kapsam 2 - Elektrik
        scope2_data = [
            {'facility_name': 'ASANSÖR D2', 'electricity_kwh': Decimal('7955030')},
            {'facility_name': 'DÖKÜMHANE D3', 'electricity_kwh': Decimal('7118068')},
            {'facility_name': 'FRENBU', 'electricity_kwh': Decimal('753468')},
        ]
        
        for data in scope2_data:
            Scope2Data.objects.update_or_create(
                firm=self.firm,
                facility_name=data['facility_name'],
                period_year=self.year,
                period_month=self.month or 1,
                defaults={
                    'electricity_kwh': data['electricity_kwh'],
                    'created_by': self.user
                }
            )
        
        # Kapsam 4 - Satın Alınan Malzemeler
        material_mappings = {
            'SİLİSLİ SAC': 'Çelik',
            'PİK DÖKÜM': 'Dökme Demir',
            'Sfero Karbon Verici': 'Dökme Demir',
            'Magnezyum': 'Magnezyum',
        }
        
        scope4_data = [
            {'product_name': 'SİLİSLİ SAC', 'quantity': Decimal('810000')},
            {'product_name': 'PİK DÖKÜM', 'quantity': Decimal('439000')},
            {'product_name': 'Sfero Karbon Verici', 'quantity': Decimal('138000')},
            {'product_name': 'Magnezyum', 'quantity': Decimal('2000')},
        ]
        
        for data in scope4_data:
            material_type = material_mappings.get(data['product_name'], 'Çelik')
            emission_factor = EmissionFactor.objects.get(
                name=material_type,
                factor_type='MATERIAL'
            )
            
            Scope4Data.objects.update_or_create(
                firm=self.firm,
                product_name=data['product_name'],
                period_year=self.year,
                period_month=self.month or 1,
                defaults={
                    'product_category': 'Hammadde',
                    'quantity': data['quantity'],
                    'unit': 'kg',
                    'emission_factor': emission_factor.ef_co2,
                    'emission_factor_source': emission_factor.source,
                    'created_by': self.user
                }
            )
        
        print("✅ Örnek veriler başarıyla yüklendi!")
    
    @transaction.atomic
    def create_report(self):
        """Rapor oluştur ve hesaplamaları yap"""
        
        # Rapor tarihlerini belirle
        if self.month:
            report_date = date(self.year, self.month, 28)
            period_start = date(self.year, self.month, 1)
            period_end = date(self.year, self.month, 28)
        else:
            report_date = date(self.year, 12, 31)
            period_start = date(self.year, 1, 1)
            period_end = date(self.year, 12, 31)
        
        # Rapor oluştur veya güncelle
        self.report, created = Report.objects.update_or_create(
            firm=self.firm,
            report_year=self.year,
            report_month=self.month,
            defaults={
                'report_date': report_date,
                'report_period_start': period_start,
                'report_period_end': period_end,
                'generated_by': self.user,
                'status': 'DRAFT'
            }
        )
        
        # Kapsam toplamlarını hesapla
        self.report.calculate_totals()
        self.report.save()
        
        # Ürünlere dağıt
        self.allocate_to_products()
        
        # Rapor durumunu güncelle
        self.report.status = 'COMPLETED'
        self.report.save()
        
        return self.report
    
    def allocate_to_products(self):
        """Excel'deki ürün dağılımını uygula"""
        
        products_data = [
            {
                'product_name': 'ASANSÖR MOTORU',
                'annual_production': 60860,
                'annual_weight_kg': Decimal('21301000')
            },
            {
                'product_name': 'KASNAK',
                'annual_production': 41184,
                'annual_weight_kg': Decimal('2059200')
            },
            {
                'product_name': 'PORYA',
                'annual_production': 166000,
                'annual_weight_kg': Decimal('7470000')
            },
            {
                'product_name': 'KAMPANA',
                'annual_production': 225806,
                'annual_weight_kg': Decimal('13548360')
            },
            {
                'product_name': 'FREN DİSKİ',
                'annual_production': 374828,
                'annual_weight_kg': Decimal('15742776')
            }
        ]
        
        # Toplam ağırlığı hesapla
        total_weight = sum(p['annual_weight_kg'] for p in products_data)
        
        for product in products_data:
            allocation = ProductCarbonAllocation.objects.update_or_create(
                firm=self.firm,
                product_name=product['product_name'],
                period_year=self.year,
                defaults={
                    'annual_production': product['annual_production'],
                    'annual_weight_kg': product['annual_weight_kg'],
                    'report': self.report
                }
            )[0]
            
            # Dağılımı hesapla
            allocation.calculate_allocation(total_weight, self.report.total_co2e)
            allocation.save()
    
    def generate_excel_report(self):
        """Excel formatında rapor oluştur"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        
        # KAPSAM 1 Sheet'i
        ws1 = wb.active
        ws1.title = "KAPSAM 1"
        ws1.append(['KAPSAM 1: Doğrudan Sera Gazı Emisyonları'])
        ws1.append([])
        ws1.append(['Lokasyon', 'Yakıt', 'Tüketim', 'Birim', 'CO2 (ton)', 'CH4 (ton)', 'N2O (ton)', 'CO2e (ton)'])
        
        scope1_data = Scope1Data.objects.filter(
            firm=self.firm,
            period_year=self.year,
            period_month=self.month or 1
        )
        
        for data in scope1_data:
            ws1.append([
                data.location,
                data.fuel_name,
                float(data.consumption_value),
                data.consumption_unit,
                float(data.co2_emission),
                float(data.ch4_emission),
                float(data.n2o_emission),
                float(data.total_co2e)
            ])
        
        # KAPSAM 2 Sheet'i
        ws2 = wb.create_sheet("KAPSAM 2")
        ws2.append(['KAPSAM 2: İthal Edilen Enerji'])
        ws2.append([])
        ws2.append(['Tesis', 'Elektrik (kWh)', 'Elektrik (MWh)', 'EF (tCO2/MWh)', 'CO2e (ton)'])
        
        scope2_data = Scope2Data.objects.filter(
            firm=self.firm,
            period_year=self.year,
            period_month=self.month or 1
        )
        
        for data in scope2_data:
            ws2.append([
                data.facility_name,
                float(data.electricity_kwh),
                float(data.electricity_mwh),
                float(data.emission_factor),
                float(data.total_co2e)
            ])
        
        # TOPLAM Sheet'i
        ws_total = wb.create_sheet("TOPLAM KARBON MİKTARI")
        ws_total.append(['KAPSAM', 'AÇIKLAMA', 'TOPLAM (tCO2e)'])
        ws_total.append(['KAPSAM 1', 'Doğrudan Emisyonlar', float(self.report.scope1_total)])
        ws_total.append(['KAPSAM 2', 'İthal Edilen Enerji', float(self.report.scope2_total)])
        ws_total.append(['KAPSAM 3', 'Ulaşım', float(self.report.scope3_total)])
        ws_total.append(['KAPSAM 4', 'Satın Alınan Ürünler', float(self.report.scope4_total)])
        ws_total.append([])
        ws_total.append(['TOPLAM', '', float(self.report.total_co2e)])
        
        # ÜRÜN HESABI Sheet'i
        ws_product = wb.create_sheet("ÜRÜN HESABI")
        ws_product.append(['ÜRÜN', 'Adet/Yıl', 'KG/YIL', '%', 'TOPLAM tCO2e', 'ÜRÜN tCO2e', '1 ADET ÜRÜN'])
        
        allocations = ProductCarbonAllocation.objects.filter(
            firm=self.firm,
            period_year=self.year
        )
        
        for allocation in allocations:
            ws_product.append([
                allocation.product_name,
                allocation.annual_production,
                float(allocation.annual_weight_kg),
                float(allocation.weight_percentage),
                float(allocation.allocated_co2e),
                float(allocation.allocated_co2e),
                float(allocation.co2e_per_unit)
            ])
        
        # BytesIO'ya kaydet
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file